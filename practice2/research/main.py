from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.cell import get_column_letter

start_x = 2
start_y = 5
offset = 10
main_mass = {}


def foo(filename='u.txt', codename='file1'):
    parses = open(filename, 'r').read().splitlines()

    d_current = 0
    for parse in parses:
        if parse.startswith("d: "):
            d_current = int(parse.split("d: ")[1])
        elif parse.startswith("N: "):
            if d_current not in main_mass:
                main_mass[d_current] = {}
            parse = parse.split("N: ")[1].split(" Time: ")
            N = int(parse[0])
            parse = parse[1].split(" count: ")
            Time = parse[0]
            # Time = format(float(Time), '.7f').replace('.',',')
            count = int(parse[1])
            if N not in main_mass[d_current]:
                main_mass[d_current][N] = {}
            # print(codename)
            main_mass[d_current][N][codename] = {
                'time': Time,
                'count': count
            }
    return main_mass


def toFile(wb, f):
    # ws = wb.active
    for d in f:
        ws_name = "d=" + str(d)
        if ws_name not in wb.sheetnames:
            wb.create_sheet(ws_name)
        ws = wb[ws_name]
        max_codenames = 0
        set_header = 0
        for idy, N in enumerate(f[d]):
            # Generate header
            if set_header == 0:
                set_header = 1
                ws.cell(row=start_y - 1, column=start_x - 1, value=None)
                for idx, codename in enumerate(f[d][N]):
                    ws.cell(row=start_y - 1, column=start_x + idx, value=codename)
                    max_codenames += 1
            ws.cell(row=start_y + idy, column=start_x - 1, value=N)
            # print(start_y, start_x + idy, N)
            for idx, codename in enumerate(f[d][N]):
                tf = f[d][N][codename]
                _cell = ws.cell(row=start_y + idy, column=start_x + idx, value=float(tf['time']))
                _cell.number_format = '0.00E+00'

        values = Reference(ws, min_col=start_x, max_col=start_x + max_codenames - 1, min_row=start_y-1,
                           max_row=start_y - 1 + len(f[d]))
        cats = Reference(ws, min_col=start_x-1, max_col=start_x-1, min_row=start_y,
                                              max_row=start_y - 1 + len(f[d]))
        # titles = Reference(ws, min_col=start_x+1, max_col=start_x + max_codenames - 1, min_row=start_y - 1,
        #                    max_row=start_y - 1)
        # values = Reference(ws, min_col=start_x-1, min_row=start_y,
        #                    max_col=start_x+len(f[d])-1, max_row=start_y+max_codenames)
        # values = Reference(ws, min_col=start_x, min_row=start_y,
        #                    max_col=start_x + len(f[d]) - 1, max_row=start_y + max_codenames)
        # # categor = Reference(ws, min_col=start_x, min_row=start_y,
        # #                    max_col=start_x+len(f[d])-1, max_row=start_y)
        # # categor = Reference(ws, min_col=start_x-1, min_row=start_y+1,
        # #                     max_col=start_x-1, max_row=start_y+max_codenames)
        # titles = Reference(ws, min_col=start_x-1, min_row=start_y+1, max_row=start_y+1+max_codenames)
        chart = LineChart()
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.grouping = 'stacked'
        #
        # # chart.titl
        ws.add_chart(chart, get_column_letter(start_x + max_codenames + offset) + str(start_y))


if __name__ == '__main__':
    wb = Workbook()

    foo('default.txt', 'default')
    foo('svyaz.txt', 'sv')
    toFile(wb, main_mass)
    # Save the file
    wb.save("sample.xlsx")
